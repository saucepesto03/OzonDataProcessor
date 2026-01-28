import os
import pandas as pd
import warnings
from pathlib import Path
import sys
from openpyxl import load_workbook
from datetime import datetime

# Отключаем предупреждения
warnings.filterwarnings('ignore')

# ========== ИСПРАВЛЕНА КРИТИЧЕСКАЯ ОШИБКА ==========
# Определяем правильную рабочую директорию для exe
if getattr(sys, 'frozen', False):
    # Запущено из .exe
    BASE_DIR = Path(sys.executable).parent
else:
    # Запущено из .py
    BASE_DIR = Path(__file__).parent
# ==================================================

def find_latest_files(folder_path, extension='.xlsx', count=1):
    """Находит самые свежие файлы в папке по дате изменения"""
    folder = Path(folder_path)
    files = list(folder.glob(f'*{extension}'))
    if not files:
        return [] if count > 1 else None
    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if count == 1:
        return str(files[0])
    return [str(f) for f in files[:count]]

def read_excel_with_error_handling(file_path, sheet_name=0, header=None):
    """Чтение Excel файла с обработкой ошибок"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header, engine='openpyxl')
        return df
    except Exception as e:
        print(f"Ошибка при чтении файла {Path(file_path).name}: {e}")
        return None

def create_backup_filename(original_path):
    """Создает имя для резервной копии файла"""
    original_path = Path(original_path)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_name = f"{original_path.stem}_обновленный_{timestamp}{original_path.suffix}"
    return original_path.parent / new_name

def process_analytics_files(analytics_files):
    """Корректная обработка одного или двух analytics файлов"""
    all_analytics_data = {}

    for file_idx, file_path in enumerate(analytics_files):
        print(f"   Чтение файла {file_idx+1}: {Path(file_path).name}")
        df_analytics = read_excel_with_error_handling(file_path, header=None)

        if df_analytics is None:
            continue

        for idx, row in df_analytics.iterrows():
            if idx >= 13:  # с 14-й строки
                sku = row[7] if len(row) > 7 else None
                if pd.isna(sku):
                    continue

                sku = str(sku).strip()

                avg_price = row[63] if len(row) > 63 else None  # BL
                total_ddr = row[69] if len(row) > 69 else None  # BR

                # ===== КЛЮЧЕВОЕ ИСПРАВЛЕНИЕ =====
                if sku not in all_analytics_data:
                    all_analytics_data[sku] = {
                        'bl': avg_price,
                        'br': total_ddr
                    }
                else:
                    if avg_price is not None:
                        all_analytics_data[sku]['bl'] = avg_price
                    if total_ddr is not None:
                        all_analytics_data[sku]['br'] = total_ddr
                # =================================

        print(f"   Обработано строк из файла {file_idx+1}")

    return all_analytics_data

def update_unit_file():
    # Пути к папкам (ОТНОСИТЕЛЬНЫЕ ОТ РАБОЧЕЙ ДИРЕКТОРИИ)
    base_dir = BASE_DIR
    mark_folder = base_dir / "MARK_ozon_report"
    analytics_folder = base_dir / "analytics_report"
    dimensions_folder = base_dir / "ozon_dimensions"
    prices_folder = base_dir / "prices_with_co-investment"
    unit_folder = base_dir / "unit_folder"
    
    # Проверяем существование папок
    print(f"Рабочая директория: {base_dir}")
    print(f"Проверяем папки:")
    print(f"  MARK_ozon_report: {'✓' if mark_folder.exists() else '✗'}")
    print(f"  analytics_report: {'✓' if analytics_folder.exists() else '✗'}")
    print(f"  ozon_dimensions: {'✓' if dimensions_folder.exists() else '✗'}")
    print(f"  prices_with_co-investment: {'✓' if prices_folder.exists() else '✗'}")
    print(f"  unit_folder: {'✓' if unit_folder.exists() else '✗'}")
    
    # Находим самые свежие файлы
    mark_file = find_latest_files(mark_folder)
    # Изменение: ищем ДВА последних файла в analytics_report
    analytics_files = find_latest_files(analytics_folder, count=2)
    dimensions_file = find_latest_files(dimensions_folder)
    prices_file = find_latest_files(prices_folder)
    unit_file = find_latest_files(unit_folder, '.xlsm')
    
    if not unit_file:
        unit_file = find_latest_files(unit_folder, '.xlsx')
    
    # Проверяем файлы
    if not mark_file:
        print(f"\n❌ Не найден файл в папке MARK_ozon_report")
    if not analytics_files:
        print(f"\n❌ Не найдены файлы в папке analytics_report")
    if not dimensions_file:
        print(f"\n❌ Не найден файл в папке ozon_dimensions")
    if not prices_file:
        print(f"\n❌ Не найден файл в папке prices_with_co-investment")
    if not unit_file:
        print(f"\n❌ Не найден файл в папке unit_folder")
    
    if not all([mark_file, analytics_files, dimensions_file, prices_file, unit_file]):
        print("\nПодробная информация о папках:")
        
        # Покажем что есть в папках
        for folder, name in [
            (mark_folder, "MARK_ozon_report"),
            (analytics_folder, "analytics_report"),
            (dimensions_folder, "ozon_dimensions"),
            (prices_folder, "prices_with_co-investment"),
            (unit_folder, "unit_folder")
        ]:
            if folder.exists():
                files = list(folder.glob('*'))
                print(f"\n  {name}: {len(files)} файлов")
                for f in files:
                    print(f"    - {f.name} ({datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')})")
            else:
                print(f"\n  {name}: папка не существует")
        
        return False
    
    print(f"\nИспользуемые файлы:")
    print(f"  Таблица A (MARK): {Path(mark_file).name}")
    print(f"  Таблица B (analytics):")
    for i, file_path in enumerate(analytics_files):
        print(f"    Файл {i+1}: {Path(file_path).name}")
    print(f"  Таблица C (dimensions): {Path(dimensions_file).name}")
    print(f"  Таблица D (prices): {Path(prices_file).name}")
    print(f"  Таблица F (unit): {Path(unit_file).name}")
    
    try:
        # ============================================
        # 1. Чтение таблицы A (MARK_ozon_report)
        # ============================================
        print("\n1. Чтение таблицы A (MARK_ozon_report)...")
        df_mark = read_excel_with_error_handling(mark_file, header=None)
        if df_mark is None:
            return False

        mark_data = {}
        for idx, row in df_mark.iterrows():
            if idx >= 9:  # С 10-й строки (индекс 9)
                sku = row[3] if len(row) > 3 else None  # ← СТОЛБЕЦ D
                if pd.notna(sku):
                    sku = str(sku).strip()
                    col_a_value = row[0] if 0 in row else None  # Столбец A
                    cost_value = row[15] if len(row) > 15 else None  # Столбец P
                    mark_data[sku] = {
                        'col_a': col_a_value,
                        'col_p': cost_value
                    }

        print(f"   Найдено {len(mark_data)} SKU в таблице A")

        
        # ============================================
        # 2. Чтение таблицы B (analytics_report) - ОБНОВЛЁННЫЙ БЛОК
        # ============================================
        print("2. Чтение таблицы B (analytics_report)...")
        analytics_data = process_analytics_files(analytics_files)
        
        if not analytics_data:
            print("   ⚠ Не удалось получить данные из файлов analytics")
        
        print(f"   Всего найдено {len(analytics_data)} уникальных SKU в таблице B")
        
        # ============================================
        # 3. Чтение таблицы C (ozon_dimensions)
        # ============================================
        print("3. Чтение таблицы C (ozon_dimensions)...")
        df_dimensions = read_excel_with_error_handling(dimensions_file, header=None)
        if df_dimensions is None:
            return False
            
        dimensions_data = {}
        for idx, row in df_dimensions.iterrows():
            if idx >= 1:  # Со 2-й строки (индекс 1)
                sku = row[0] if 0 in row else None  # Столбец A
                if pd.notna(sku):
                    sku = str(sku).strip()
                    length = row[5] if len(row) > 5 else None  # Столбец F (Длина)
                    width = row[3] if len(row) > 3 else None   # Столбец D (Ширина)
                    height = row[4] if len(row) > 4 else None  # Столбец E (Высота)
                    dimensions_data[sku] = {
                        'length': length,
                        'width': width,
                        'height': height
                    }
        print(f"   Найдено {len(dimensions_data)} SKU в таблице C")
        
        # ============================================
        # 4. Чтение таблицы D (prices_with_co-investment)
        # ============================================
        print("4. Чтение таблицы D (prices_with_co-investment)...")
        df_prices = read_excel_with_error_handling(prices_file, header=None)
        if df_prices is None:
            return False
            
        prices_data = {}
        for idx, row in df_prices.iterrows():
            if idx >= 1:  # Со 2-й строки (индекс 1)
                sku = row[0] if 0 in row else None  # Столбец A
                if pd.notna(sku):
                    sku = str(sku).strip()
                    price = row[2] if len(row) > 2 else None  # Столбец C
                    prices_data[sku] = price
        print(f"   Найдено {len(prices_data)} SKU в таблице D")
        
        # ============================================
        # 5. Создание обновлённой копии таблицы F (Unit.xlsm)
        # ============================================
        print("\n5. Создание обновлённой копии таблицы F...")
        
        # Создаем имя для обновлённой копии
        new_unit_file = create_backup_filename(unit_file)
        
        # Загружаем оригинальный файл с сохранением макросов
        wb = load_workbook(filename=unit_file, keep_vba=True)
        ws = wb.active
        
        updated_count = 0
        total_rows = 0
        
        # Проходим по строкам начиная со второй
        for row in range(2, ws.max_row + 1):
            sku_cell = ws.cell(row=row, column=1)  # Колонка A
            sku = sku_cell.value

            if sku is not None:
                # ===== ИСПРАВЛЕНИЕ =====
                if isinstance(sku, float):
                    sku_str = str(int(sku))
                else:
                    sku_str = str(sku).strip()
                # =======================

                total_rows += 1
                updated = False

                # Обновляем из таблицы A
                if sku_str in mark_data:
                    col_a_value = mark_data[sku_str].get('col_a')
                    if col_a_value is not None:
                        ws.cell(row=row, column=2, value=col_a_value)
                        updated = True

                    col_p_value = mark_data[sku_str].get('col_p')
                    if col_p_value is not None:
                        ws.cell(row=row, column=3, value=col_p_value)
                        updated = True

                # Обновляем из таблицы C
                if sku_str in dimensions_data:
                    dim = dimensions_data[sku_str]
                    if dim.get('length') is not None:
                        ws.cell(row=row, column=5, value=dim['length'])
                        updated = True
                    if dim.get('width') is not None:
                        ws.cell(row=row, column=6, value=dim['width'])
                        updated = True
                    if dim.get('height') is not None:
                        ws.cell(row=row, column=7, value=dim['height'])
                        updated = True

                # Обновляем из таблицы B
                if sku_str in analytics_data:
                    analytics = analytics_data[sku_str]
                    if analytics.get('bl') is not None:
                        ws.cell(row=row, column=15, value=analytics['bl'])
                        updated = True
                    if analytics.get('br') is not None:
                        ws.cell(row=row, column=34, value=analytics['br'])
                        updated = True

                # Обновляем из таблицы D
                if sku_str in prices_data:
                    price = prices_data[sku_str]
                    if price is not None:
                        ws.cell(row=row, column=37, value=price)
                        updated = True

                if updated:
                    updated_count += 1

        
        # Сохраняем обновлённую копию
        wb.save(new_unit_file)
        print(f"   Создана обновлённая копия: {new_unit_file.name}")
        print(f"   Оригинальный файл остался без изменений: {Path(unit_file).name}")
        print(f"   Обновлено {updated_count} строк из {total_rows} всего")
        
        # Выводим сводку
        print("\n" + "=" * 60)
        print("СВОДКА:")
        print("=" * 60)
        print(f"Таблица A (MARK): {len(mark_data)} SKU")
        print(f"Таблица B (analytics): {len(analytics_data)} уникальных SKU из {len(analytics_files)} файлов")
        print(f"Таблица C (dimensions): {len(dimensions_data)} SKU")
        print(f"Таблица D (prices): {len(prices_data)} SKU")
        print(f"Таблица F (unit): {total_rows} строк с SKU, {updated_count} обновлено")
        print(f"\nФайлы:")
        print(f"  Оригинал: {Path(unit_file).name}")
        print(f"  Обновлённая копия: {new_unit_file.name}")
        print("=" * 60)
        
        return True
        
    except Exception as e:
        print(f"\nОшибка при обновлении файла: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("=" * 60)
    print("СКРИПТ ОБНОВЛЕНИЯ ТАБЛИЦЫ UNIT")
    print("=" * 60)
    print("Создаётся обновлённая копия файла, оригинал остаётся без изменений.")
    print("Внимание: Обрабатываются 2 последних файла в папке analytics_report")
    print("=" * 60)
    
    success = update_unit_file()
    
    if success:
        print("\n✓ ОБНОВЛЕНИЕ ЗАВЕРШЕНО УСПЕШНО!")
        print("  Обновлённая копия создана в папке unit_folder")
    else:
        print("\n✗ ОБНОВЛЕНИЕ ЗАВЕРШИЛОСЬ С ОШИБКАМИ!")
    
    print("=" * 60)
    
    # Ожидание нажатия Enter перед закрытием (если запущено из exe)
    if getattr(sys, 'frozen', False):
        input("\nНажмите Enter для выхода...")